from __future__ import annotations

from dataclasses import dataclass
from pathlib import Path
from typing import Callable

from openpyxl import Workbook, load_workbook
from openpyxl.utils.cell import coordinate_to_tuple
from PIL import Image

from ocr_models import TemplateField


@dataclass
class ExportSettings:
    sheet_name: str = "OCR"
    write_mode: str = "上書き"
    start_row: int = 1
    start_col: int = 1
    include_filename: bool = True
    include_header: bool = True


@dataclass
class ExportError:
    image: str
    field: str
    error: str


@dataclass
class ExportResult:
    total_images: int
    errors: list[ExportError]


OcrCallback = Callable[[Image.Image, TemplateField], tuple[str | None, str | None]]


def validate_export_settings(
    sheet_name: str,
    write_mode: str,
    start_cell: str,
    include_filename: bool,
    include_header: bool,
) -> tuple[ExportSettings | None, str | None]:
    sheet_name = sheet_name.strip() or "OCR"
    if len(sheet_name) > 31 or any(char in sheet_name for char in "[]:*?/\\"):
        return None, "シート名は31文字以内で、次の文字は使えません: []:*?/\\"

    start_cell = start_cell.strip().upper() or "A1"
    try:
        start_row, start_col = coordinate_to_tuple(start_cell)
    except ValueError:
        return None, "開始セルは A1 形式で入力してください。"

    if write_mode not in {"上書き", "追記"}:
        write_mode = "上書き"

    return (
        ExportSettings(
            sheet_name=sheet_name,
            write_mode=write_mode,
            start_row=start_row,
            start_col=start_col,
            include_filename=include_filename,
            include_header=include_header,
        ),
        None,
    )


class ExcelExporter:
    def export(
        self,
        output_path: Path,
        image_files: list[Path],
        fields: list[TemplateField],
        settings: ExportSettings,
        ocr_callback: OcrCallback,
        progress_callback: Callable[[int, int, Path], None] | None = None,
    ) -> ExportResult:
        workbook, sheet, row_cursor = self._prepare_output_sheet(output_path, settings)
        headers = self._export_headers(fields, settings.include_filename)
        if settings.include_header and row_cursor == settings.start_row:
            self._write_excel_row(sheet, row_cursor, settings.start_col, headers)
            row_cursor += 1

        errors: list[ExportError] = []
        total = len(image_files)

        for row_index, image_path in enumerate(image_files, start=1):
            try:
                image = Image.open(image_path).convert("RGB")
            except Exception as exc:
                errors.append(ExportError(image_path.name, "", f"画像を開けませんでした: {exc}"))
                if progress_callback:
                    progress_callback(row_index, total, image_path)
                continue

            row = [image_path.name] if settings.include_filename else []
            for field in fields:
                text, error = ocr_callback(image, field)
                if text is None:
                    errors.append(ExportError(image_path.name, field.name, error or "OCRに失敗しました。"))
                    row.append("")
                    continue
                row.append(text)

            self._write_excel_row(sheet, row_cursor, settings.start_col, row)
            row_cursor += 1
            if progress_callback:
                progress_callback(row_index, total, image_path)

        self._fit_output_columns(sheet)
        self._write_error_sheet(workbook, errors)

        output_path.parent.mkdir(parents=True, exist_ok=True)
        workbook.save(output_path)
        return ExportResult(total, errors)

    def _prepare_output_sheet(self, output_path: Path, settings: ExportSettings):
        if settings.write_mode == "追記" and output_path.exists():
            workbook = load_workbook(output_path)
            sheet = workbook[settings.sheet_name] if settings.sheet_name in workbook.sheetnames else workbook.create_sheet(settings.sheet_name)
            row_cursor = max(settings.start_row, sheet.max_row + 1) if self._sheet_has_values(sheet) else settings.start_row
            return workbook, sheet, row_cursor

        workbook = Workbook()
        sheet = workbook.active
        sheet.title = settings.sheet_name
        return workbook, sheet, settings.start_row

    def _sheet_has_values(self, sheet) -> bool:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value not in (None, ""):
                    return True
        return False

    def _export_headers(self, fields: list[TemplateField], include_filename: bool) -> list[str]:
        headers = [field.name for field in fields]
        if include_filename:
            return ["画像名", *headers]
        return headers

    def _write_excel_row(self, sheet, row_index: int, start_col: int, values: list[str]) -> None:
        for offset, value in enumerate(values):
            sheet.cell(row=row_index, column=start_col + offset, value=value)

    def _fit_output_columns(self, sheet) -> None:
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 42)

    def _write_error_sheet(self, workbook, errors: list[ExportError]) -> None:
        if "Errors" in workbook.sheetnames:
            workbook.remove(workbook["Errors"])
        if not errors:
            return
        sheet = workbook.create_sheet("Errors")
        sheet.append(["画像ファイル", "項目", "エラー"])
        for error in errors:
            sheet.append([error.image, error.field, error.error])
        self._fit_output_columns(sheet)
