from __future__ import annotations

from dataclasses import dataclass, field
from pathlib import Path
from typing import Callable, cast

from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from openpyxl.utils.cell import coordinate_to_tuple
from openpyxl.worksheet.worksheet import Worksheet
from PIL import Image

from ImageOCR2Excel.persistence import atomic_save_workbook
from ImageOCR2Excel.models import (
    DEFAULT_SET_DEFINITION,
    EXPORT_LAYOUT_IMAGE_ROW,
    EXPORT_LAYOUT_OPTIONS,
    EXPORT_LAYOUT_SET,
    SetDetectionResult,
    SetDefinition,
    TemplateField,
    set_groups,
    set_validation_error,
)
from ImageOCR2Excel.operations import CancelCheck, raise_if_cancelled


@dataclass
class ExportSettings:
    sheet_name: str = "OCR"
    write_mode: str = "上書き"
    start_row: int = 1
    start_col: int = 1
    include_filename: bool = True
    include_header: bool = True
    output_layout: str = EXPORT_LAYOUT_IMAGE_ROW


@dataclass
class ExportError:
    image: str
    field: str
    error: str


@dataclass
class ExportResult:
    total_images: int
    errors: list[ExportError]
    records: list["ImageExportRecord"]
    notices: list["ExportNotice"] = field(default_factory=list)

    @property
    def failed_records(self) -> list["ImageExportRecord"]:
        return [record for record in self.records if record.errors]


@dataclass
class ImageExportRecord:
    image_path: Path
    start_row: int
    row_count: int
    errors: list[ExportError]
    notices: list["ExportNotice"] = field(default_factory=list)


@dataclass
class ExportNotice:
    image: str
    notice: str


OcrCallback = Callable[[Image.Image, TemplateField], tuple[str | None, str | None]]
SetResolver = Callable[
    [Image.Image, list[TemplateField]], SetDetectionResult
]


def resolve_export_set_definition(
    settings: ExportSettings,
    set_definition: SetDefinition | None,
) -> SetDefinition:
    if settings.output_layout == EXPORT_LAYOUT_SET and set_definition is None:
        raise ValueError(
            "セット単位で出力するには、セット定義を指定してください。"
        )
    return (set_definition or DEFAULT_SET_DEFINITION).normalized()


def validate_export_settings(
    sheet_name: str,
    write_mode: str,
    start_cell: str,
    include_filename: bool,
    include_header: bool,
    output_layout: str = EXPORT_LAYOUT_IMAGE_ROW,
) -> tuple[ExportSettings | None, str | None]:
    sheet_name = sheet_name.strip() or "OCR"
    if len(sheet_name) > 31 or any(char in sheet_name for char in "[]:*?/\\"):
        return None, "シート名は31文字以内で、次の文字は使えません: []:*?/\\"

    start_cell = start_cell.strip().upper() or "A1"
    try:
        start_row, start_col = coordinate_to_tuple(start_cell)
    except ValueError:
        return None, "開始セルは A1 形式で入力してください。"

    if write_mode not in {"上書き", "追記"}:
        write_mode = "上書き"
    if output_layout not in EXPORT_LAYOUT_OPTIONS:
        output_layout = EXPORT_LAYOUT_IMAGE_ROW

    return (
        ExportSettings(
            sheet_name=sheet_name,
            write_mode=write_mode,
            start_row=start_row,
            start_col=start_col,
            include_filename=include_filename,
            include_header=include_header,
            output_layout=output_layout,
        ),
        None,
    )


class ExcelExporter:
    def export(
        self,
        output_path: Path,
        image_files: list[Path],
        fields: list[TemplateField],
        settings: ExportSettings,
        ocr_callback: OcrCallback,
        progress_callback: Callable[[int, int, Path], None] | None = None,
        set_definition: SetDefinition | None = None,
        set_resolver: SetResolver | None = None,
        cancel_check: CancelCheck | None = None,
    ) -> ExportResult:
        raise_if_cancelled(cancel_check)
        definition = resolve_export_set_definition(settings, set_definition)
        workbook, sheet, row_cursor = self._prepare_output_sheet(output_path, settings)
        headers = self.headers(fields, settings, definition)
        if settings.include_header and row_cursor == settings.start_row:
            self._write_excel_row(sheet, row_cursor, settings.start_col, headers)
            row_cursor += 1

        rows, result = self.recognize_rows(
            image_files,
            fields,
            settings,
            ocr_callback,
            start_row=row_cursor,
            progress_callback=progress_callback,
            set_definition=definition,
            set_resolver=set_resolver,
            cancel_check=cancel_check,
        )
        for row in rows:
            self._write_excel_row(sheet, row_cursor, settings.start_col, row)
            row_cursor += 1

        raise_if_cancelled(cancel_check)
        self._fit_output_columns(sheet)
        self._write_error_sheet(workbook, result.errors)
        self._write_notice_sheet(workbook, result.notices)

        raise_if_cancelled(cancel_check)
        atomic_save_workbook(workbook, output_path)
        return result

    def recognize_rows(
        self,
        image_files: list[Path],
        fields: list[TemplateField],
        settings: ExportSettings,
        ocr_callback: OcrCallback,
        *,
        start_row: int = 1,
        progress_callback: Callable[[int, int, Path], None] | None = None,
        set_definition: SetDefinition | None = None,
        set_resolver: SetResolver | None = None,
        cancel_check: CancelCheck | None = None,
    ) -> tuple[list[list[str]], ExportResult]:
        """Recognize images into format-neutral rows for spreadsheet exporters."""
        raise_if_cancelled(cancel_check)
        definition = resolve_export_set_definition(settings, set_definition)
        if settings.output_layout == EXPORT_LAYOUT_SET and set_resolver is None:
            set_error = set_validation_error(fields, definition)
            if set_error:
                raise ValueError(set_error)

        output_rows: list[list[str]] = []
        errors: list[ExportError] = []
        notices: list[ExportNotice] = []
        records: list[ImageExportRecord] = []
        row_cursor = start_row
        total = len(image_files)

        for row_index, image_path in enumerate(image_files, start=1):
            raise_if_cancelled(cancel_check)
            rows, image_errors, image_notices = self.recognize_image_rows(
                image_path,
                fields,
                settings,
                ocr_callback,
                definition,
                set_resolver,
                cancel_check,
            )
            records.append(
                ImageExportRecord(
                    image_path,
                    row_cursor,
                    len(rows),
                    image_errors,
                    image_notices,
                )
            )
            output_rows.extend(rows)
            row_cursor += len(rows)
            errors.extend(image_errors)
            notices.extend(image_notices)
            if progress_callback:
                progress_callback(row_index, total, image_path)

        raise_if_cancelled(cancel_check)
        return output_rows, ExportResult(total, errors, records, notices)

    def retry_failed(
        self,
        output_path: Path,
        previous_result: ExportResult,
        fields: list[TemplateField],
        settings: ExportSettings,
        ocr_callback: OcrCallback,
        retry_files: list[Path] | None = None,
        progress_callback: Callable[[int, int, Path], None] | None = None,
        set_definition: SetDefinition | None = None,
        set_resolver: SetResolver | None = None,
        cancel_check: CancelCheck | None = None,
    ) -> ExportResult:
        raise_if_cancelled(cancel_check)
        definition = resolve_export_set_definition(settings, set_definition)
        if settings.output_layout == EXPORT_LAYOUT_SET and set_resolver is None:
            set_error = set_validation_error(fields, definition)
            if set_error:
                raise ValueError(set_error)
        if not output_path.exists():
            raise FileNotFoundError(f"再実行するExcelファイルがありません: {output_path}")
        workbook = load_workbook(output_path)
        if settings.sheet_name not in workbook.sheetnames:
            raise ValueError(f"再実行先のシートがありません: {settings.sheet_name}")
        sheet = workbook[settings.sheet_name]
        retry_targets = set(retry_files) if retry_files is not None else None
        retry_records = [
            record for record in previous_result.failed_records
            if retry_targets is None or record.image_path in retry_targets
        ]
        retry_paths = {record.image_path for record in retry_records}
        errors = [error for record in previous_result.records if record.image_path not in retry_paths for error in record.errors]
        notices = [notice for record in previous_result.records if record.image_path not in retry_paths for notice in record.notices]
        updated_by_path: dict[Path, ImageExportRecord] = {}
        total = len(retry_records)

        for row_index, record in enumerate(retry_records, start=1):
            raise_if_cancelled(cancel_check)
            rows, image_errors, image_notices = self.recognize_image_rows(
                record.image_path,
                fields,
                settings,
                ocr_callback,
                definition,
                set_resolver,
                cancel_check,
            )
            if len(rows) != record.row_count:
                raise ValueError("出力レイアウトが前回の一括出力から変更されています。全画像を再出力してください。")
            self._clear_record_rows(
                sheet,
                record,
                settings.start_col,
                len(self.headers(fields, settings, definition)),
            )
            for offset, row in enumerate(rows):
                self._write_excel_row(sheet, record.start_row + offset, settings.start_col, row)
            errors.extend(image_errors)
            notices.extend(image_notices)
            updated_by_path[record.image_path] = ImageExportRecord(
                record.image_path,
                record.start_row,
                record.row_count,
                image_errors,
                image_notices,
            )
            if progress_callback:
                progress_callback(row_index, total, record.image_path)

        raise_if_cancelled(cancel_check)
        self._fit_output_columns(sheet)
        self._write_error_sheet(workbook, errors)
        self._write_notice_sheet(workbook, notices)
        raise_if_cancelled(cancel_check)
        atomic_save_workbook(workbook, output_path)
        records = [updated_by_path.get(record.image_path, record) for record in previous_result.records]
        return ExportResult(previous_result.total_images, errors, records, notices)

    def recognize_image_rows(
        self,
        image_path: Path,
        fields: list[TemplateField],
        settings: ExportSettings,
        ocr_callback: OcrCallback,
        set_definition: SetDefinition | None = None,
        set_resolver: SetResolver | None = None,
        cancel_check: CancelCheck | None = None,
    ) -> tuple[list[list[str]], list[ExportError], list[ExportNotice]]:
        try:
            image = Image.open(image_path).convert("RGB")
        except Exception as exc:
            error = ExportError(image_path.name, "", f"画像を開けませんでした: {exc}")
            return self._blank_rows(
                image_path,
                fields,
                settings,
                set_definition or DEFAULT_SET_DEFINITION,
            ), [error], []

        if settings.output_layout == EXPORT_LAYOUT_SET:
            return self._recognize_set_rows(
                image_path,
                image,
                fields,
                settings,
                ocr_callback,
                set_definition or DEFAULT_SET_DEFINITION,
                set_resolver,
                cancel_check,
            )

        errors: list[ExportError] = []
        row = [image_path.name] if settings.include_filename else []
        for field in fields:
            raise_if_cancelled(cancel_check)
            text, error = ocr_callback(image, field)
            raise_if_cancelled(cancel_check)
            if text is None:
                errors.append(ExportError(image_path.name, field.name, error or "OCRに失敗しました。"))
                row.append("")
                continue
            if not text.strip():
                errors.append(ExportError(image_path.name, field.name, "OCR結果が空です。読み取り範囲または認識設定を確認してください。"))
            row.append(text)
        return [row], errors, []

    def _recognize_set_rows(
        self,
        image_path: Path,
        image: Image.Image,
        fields: list[TemplateField],
        settings: ExportSettings,
        ocr_callback: OcrCallback,
        set_definition: SetDefinition,
        set_resolver: SetResolver | None = None,
        cancel_check: CancelCheck | None = None,
    ) -> tuple[list[list[str]], list[ExportError], list[ExportNotice]]:
        errors: list[ExportError] = []
        notices: list[ExportNotice] = []
        rows: list[list[str]] = []
        allow_empty_slots: set[str] = set()
        if set_resolver is None:
            groups = set_groups(fields, set_definition)
        else:
            raise_if_cancelled(cancel_check)
            detection = set_resolver(image, fields)
            raise_if_cancelled(cancel_check)
            groups = detection.groups
            allow_empty_slots.update(detection.allow_empty_slots)
            notices.extend(
                ExportNotice(image_path.name, message)
                for message in detection.notices
            )
            if detection.review_reason:
                errors.append(
                    ExportError(image_path.name, "セット確認", detection.review_reason)
                )
                return [], errors, notices
            if not groups:
                errors.append(
                    ExportError(
                        image_path.name,
                        "セット検出",
                        detection.review_reason
                        or "完全なセットを検出できませんでした。",
                    )
                )
        for order, slots in groups:
            missing_keys = [
                column.key
                for column in set_definition.columns
                if column.key not in slots
            ]
            if missing_keys:
                labels = [
                    set_definition.slot_label(key)
                    for key in missing_keys
                ]
                errors.append(
                    ExportError(
                        image_path.name,
                        "セット検出",
                        "自動検出結果に必要な項目がありません: "
                        + "、".join(labels),
                    )
                )
                continue
            values: list[str] = []
            for column in set_definition.columns:
                raise_if_cancelled(cancel_check)
                field = slots[column.key]
                text, error = ocr_callback(image, field)
                raise_if_cancelled(cancel_check)
                failed = text is None
                if text is None:
                    errors.append(ExportError(image_path.name, field.name, error or "OCRに失敗しました。"))
                    text = ""
                if not failed and not text.strip() and column.key not in allow_empty_slots:
                    errors.append(ExportError(image_path.name, field.name, "OCR結果が空です。読み取り範囲または認識設定を確認してください。"))
                values.append(text)
            row = [str(order), *values]
            if settings.include_filename:
                row.insert(0, image_path.name)
            rows.append(row)
        return rows, errors, notices

    def _blank_rows(
        self,
        image_path: Path,
        fields: list[TemplateField],
        settings: ExportSettings,
        set_definition: SetDefinition,
    ) -> list[list[str]]:
        if settings.output_layout == EXPORT_LAYOUT_SET:
            rows = []
            for order, _slots in set_groups(fields, set_definition):
                row = [str(order), *([""] * len(set_definition.columns))]
                if settings.include_filename:
                    row.insert(0, image_path.name)
                rows.append(row)
            return rows
        return [[*([image_path.name] if settings.include_filename else []), *([""] * len(fields))]]

    def headers(
        self,
        fields: list[TemplateField],
        settings: ExportSettings,
        set_definition: SetDefinition,
    ) -> list[str]:
        if settings.output_layout == EXPORT_LAYOUT_SET:
            return self._set_headers(settings.include_filename, set_definition)
        return self._export_headers(fields, settings.include_filename)

    def _clear_record_rows(self, sheet, record: ImageExportRecord, start_col: int, column_count: int) -> None:
        for row in range(record.start_row, record.start_row + record.row_count):
            for column in range(start_col, start_col + column_count):
                sheet.cell(row=row, column=column).value = None

    def _prepare_output_sheet(self, output_path: Path, settings: ExportSettings):
        if settings.write_mode == "追記" and output_path.exists():
            workbook = load_workbook(output_path)
            sheet = cast(
                Worksheet,
                workbook[settings.sheet_name]
                if settings.sheet_name in workbook.sheetnames
                else workbook.create_sheet(settings.sheet_name),
            )
            row_cursor = max(settings.start_row, sheet.max_row + 1) if self._sheet_has_values(sheet) else settings.start_row
            return workbook, sheet, row_cursor

        workbook = Workbook()
        sheet = workbook.active
        if sheet is None:
            sheet = workbook.create_sheet()
        sheet.title = settings.sheet_name
        return workbook, sheet, settings.start_row

    def _sheet_has_values(self, sheet) -> bool:
        for row in sheet.iter_rows():
            for cell in row:
                if cell.value not in (None, ""):
                    return True
        return False

    def _export_headers(self, fields: list[TemplateField], include_filename: bool) -> list[str]:
        headers = [field.name for field in fields]
        if include_filename:
            return ["画像名", *headers]
        return headers

    def _set_headers(self, include_filename: bool, definition: SetDefinition) -> list[str]:
        headers = [definition.order_label, *(column.label for column in definition.columns)]
        if include_filename:
            return ["画像名", *headers]
        return headers

    def _write_excel_row(self, sheet, row_index: int, start_col: int, values: list[str]) -> None:
        max_lines = 1
        for offset, value in enumerate(values):
            cell = sheet.cell(row=row_index, column=start_col + offset, value=value)
            line_count = str(value).count("\n") + 1
            max_lines = max(max_lines, line_count)
            if line_count > 1:
                cell.alignment = Alignment(wrap_text=True, vertical="top")
        if max_lines > 1:
            sheet.row_dimensions[row_index].height = max(15, 15 * max_lines)

    def _fit_output_columns(self, sheet) -> None:
        for column_cells in sheet.columns:
            max_length = max(len(str(cell.value or "")) for cell in column_cells)
            sheet.column_dimensions[column_cells[0].column_letter].width = min(max(max_length + 2, 12), 42)

    def _write_error_sheet(self, workbook, errors: list[ExportError]) -> None:
        if "Errors" in workbook.sheetnames:
            workbook.remove(workbook["Errors"])
        if not errors:
            return
        sheet = workbook.create_sheet("Errors")
        sheet.append(["画像ファイル", "項目", "エラー"])
        for error in errors:
            sheet.append([error.image, error.field, error.error])
        self._fit_output_columns(sheet)

    def _write_notice_sheet(self, workbook, notices: list[ExportNotice]) -> None:
        if "Notices" in workbook.sheetnames:
            workbook.remove(workbook["Notices"])
        if not notices:
            return
        sheet = workbook.create_sheet("Notices")
        sheet.append(["画像ファイル", "処理通知"])
        for notice in notices:
            sheet.append([notice.image, notice.notice])
        self._fit_output_columns(sheet)

