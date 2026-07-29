from __future__ import annotations

import json
import tempfile
import unittest
from pathlib import Path

from openpyxl import load_workbook
from PIL import Image

from ImageOCR2Excel.export.excel import ExcelExporter, ExportSettings
from ImageOCR2Excel.models import (
    DEFAULT_SET_DEFINITION,
    EXPORT_LAYOUT_IMAGE_ROW,
    TEMPLATE_VERSION,
    CoordinateSettings,
    TemplateField,
)
from ImageOCR2Excel.profiles import get_profile, registered_profiles
from ImageOCR2Excel.templates import (
    TemplateValidationError,
    build_template_data,
    load_template,
    save_template,
)


def valid_template() -> dict:
    return build_template_data(
        fields=[TemplateField("値", 0, 0, 80, 30)],
        lang="jpn",
        ocr_backend="paddle",
        output_settings={
            "sheet_name": "OCR",
            "write_mode": "上書き",
            "start_cell": "A1",
            "include_filename": True,
            "include_header": True,
            "output_layout": EXPORT_LAYOUT_IMAGE_ROW,
        },
        coordinate_settings=CoordinateSettings(),
        set_definition=DEFAULT_SET_DEFINITION,
        profile_id="generic",
    )


class FoundationTests(unittest.TestCase):
    def test_generic_profile_is_the_default_image_ocr_profile(self) -> None:
        profile = get_profile("generic")

        self.assertIn(profile, registered_profiles())
        self.assertEqual(profile.profile_id, "generic")
        self.assertEqual(profile.default_backend, "paddle")
        self.assertEqual(profile.default_lang, "jpn")
        self.assertIsNone(profile.auto_detection)
        self.assertEqual(profile.default_set_definition, DEFAULT_SET_DEFINITION)

    def test_template_version_one_starts_without_legacy_tesseract_state(self) -> None:
        data = valid_template()

        self.assertEqual(TEMPLATE_VERSION, 1)
        self.assertEqual(data["version"], 1)
        self.assertEqual(data["profile_id"], "generic")
        self.assertEqual(data["ocr_backend"], "paddle")
        self.assertNotIn("tesseract_path", data)

    def test_template_round_trip_rejects_non_paddle_backend(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "template.json"
            data = valid_template()

            save_template(path, data)
            loaded = load_template(path, expected_profile_id="generic")
            self.assertEqual(loaded["profile_id"], "generic")

            data["ocr_backend"] = "tesseract"
            path.write_text(json.dumps(data), encoding="utf-8")
            with self.assertRaisesRegex(TemplateValidationError, "paddle"):
                load_template(path)

    def test_old_mvp_template_version_is_not_migrated(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            path = Path(temp_dir) / "old-template.json"
            path.write_text(
                json.dumps(
                    {
                        "format": "image-ocr-to-excel-template",
                        "version": 3,
                        "template_name": "old",
                        "lang": "jpn+eng",
                        "sample_image": "",
                        "output_settings": {},
                        "fields": [
                            {
                                "name": "値",
                                "x1": 0,
                                "y1": 0,
                                "x2": 10,
                                "y2": 10,
                            }
                        ],
                    }
                ),
                encoding="utf-8",
            )

            with self.assertRaisesRegex(TemplateValidationError, "version 1"):
                load_template(path)

    def test_excel_exporter_keeps_generic_image_row_output(self) -> None:
        with tempfile.TemporaryDirectory() as temp_dir:
            root = Path(temp_dir)
            image_path = root / "sample.png"
            output_path = root / "output.xlsx"
            Image.new("RGB", (20, 20), "white").save(image_path)

            result = ExcelExporter().export(
                output_path,
                [image_path],
                [TemplateField("値", 0, 0, 20, 20)],
                ExportSettings(),
                lambda _image, _field: ("認識結果", None),
            )

            sheet = load_workbook(output_path)["OCR"]
            self.assertEqual(result.total_images, 1)
            self.assertEqual(sheet.cell(1, 1).value, "画像名")
            self.assertEqual(sheet.cell(1, 2).value, "値")
            self.assertEqual(sheet.cell(2, 1).value, "sample.png")
            self.assertEqual(sheet.cell(2, 2).value, "認識結果")


if __name__ == "__main__":
    unittest.main()
